#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Модуль формирования описи и итогового пакета документов

Реализует требования FR-3.11, FR-3.12 из расширенного ТЗ:
- Формирование структурированной описи документов
- Экспорт описи в различные форматы
- Создание структуры каталогов пакета
- Упаковка в ZIP-архив
"""

import os
import json
import shutil
import zipfile
import csv
import logging
from typing import List, Dict, Optional, Any
from pathlib import Path
from datetime import datetime
import hashlib

logger = logging.getLogger(__name__)


class PackageManifest:
    """
    Формирование описи и итогового пакета документов (FR-3.11, FR-3.12)

    Обеспечивает:
    - Создание структурированной описи всех документов
    - Экспорт в JSON, CSV, Excel (xlsx)
    - Формирование структуры каталогов пакета
    - Упаковку в ZIP-архив
    """

    # Структура каталогов пакета
    PACKAGE_STRUCTURE = {
        "forms": "Формы",
        "templates": "Документы_из_каталога",
        "user_docs": "Документы_пользователя",
        "requirements": "Требования",
    }

    # Статусы готовности документов
    STATUS_NOT_PREPARED = "not_prepared"
    STATUS_FROM_TEMPLATE = "from_template"
    STATUS_PROVIDED = "provided"
    STATUS_MISSING = "missing"

    def __init__(self, output_dir: str = "./output"):
        """
        Args:
            output_dir: Директория для вывода пакетов
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_manifest(
        self,
        documents: List[Dict],
        requirements: List[Dict],
        procurement_info: Optional[Dict] = None
    ) -> Dict[str, Any]:
        """
        Создание структурированной описи документов

        Args:
            documents: Список документов с метаданными
            requirements: Список требований из анализа КД
            procurement_info: Информация о закупке

        Returns:
            Структурированная опись
        """
        manifest_id = f"MAN-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # Создаем записи описи
        manifest_items = []
        position = 0

        for req in requirements:
            position += 1
            req_id = req.get("id", f"REQ-{position:03d}")
            req_name = req.get("name", "Неизвестный документ")
            mandatory = req.get("mandatory", True)

            # Ищем соответствующий документ
            matched_doc = self._find_matching_document(req, documents)

            # Определяем статус и источник
            if matched_doc:
                status = matched_doc.get("source_type", self.STATUS_PROVIDED)
                source_path = matched_doc.get("file_path", "")
                target_path = matched_doc.get("target_path", "")
            else:
                status = self.STATUS_MISSING if mandatory else self.STATUS_NOT_PREPARED
                source_path = ""
                target_path = ""

            item = {
                "position": position,
                "document_id": req_id,
                "document_name": req_name,
                "type": req.get("category", "other"),
                "mandatory": mandatory,
                "completion_status": status,
                "source_type": self._determine_source_type(matched_doc),
                "source_path": source_path,
                "target_path": target_path,
                "linked_requirements": [req_id],
                "notes": self._generate_notes(req, matched_doc),
                "format": req.get("format", ""),
                "validity_requirements": req.get("validity_requirements", ""),
            }

            manifest_items.append(item)

        # Расчет метрик
        metrics = self._calculate_metrics(manifest_items)

        manifest = {
            "manifest_id": manifest_id,
            "procurement_info": procurement_info or {},
            "created_at": datetime.now().isoformat(),
            "items": manifest_items,
            "metrics": metrics,
            "total_documents": len(manifest_items),
        }

        logger.info(f"Опись создана: {manifest_id}, документов: {len(manifest_items)}")
        return manifest

    def _find_matching_document(self, requirement: Dict, documents: List[Dict]) -> Optional[Dict]:
        """Поиск документа, соответствующего требованию"""
        req_name = requirement.get("name", "").lower()

        for doc in documents:
            doc_name = doc.get("name", "").lower()
            if req_name in doc_name or doc_name in req_name:
                return doc

        return None

    def _determine_source_type(self, document: Optional[Dict]) -> str:
        """Определение источника документа"""
        if not document:
            return "not_found"

        source = document.get("source_type", "")
        if source:
            return source

        if document.get("from_template"):
            return "template_library"
        if document.get("user_uploaded"):
            return "user_provided"

        return "from_KD"

    def _generate_notes(self, requirement: Dict, document: Optional[Dict]) -> str:
        """Генерация примечаний для документа"""
        notes = []

        if not document:
            if requirement.get("mandatory"):
                notes.append("ТРЕБУЕТСЯ: обязательный документ отсутствует")
            else:
                notes.append("Опциональный документ не предоставлен")
        else:
            if document.get("from_template"):
                notes.append("Скопирован из каталога типовых документов")
            if document.get("expiring_soon"):
                notes.append("Внимание: срок действия скоро истекает")

        return "; ".join(notes) if notes else ""

    def _calculate_metrics(self, items: List[Dict]) -> Dict[str, Any]:
        """Расчет метрик готовности пакета"""
        total = len(items)
        mandatory = [i for i in items if i.get("mandatory")]
        provided = [i for i in items if i.get("completion_status") in [self.STATUS_PROVIDED, self.STATUS_FROM_TEMPLATE]]
        mandatory_provided = [i for i in mandatory if i.get("completion_status") in [self.STATUS_PROVIDED, self.STATUS_FROM_TEMPLATE]]

        return {
            "total_documents": total,
            "mandatory_documents": len(mandatory),
            "provided_documents": len(provided),
            "mandatory_provided": len(mandatory_provided),
            "completeness_percentage": round(
                len(mandatory_provided) / len(mandatory) * 100, 1
            ) if mandatory else 100,
            "missing_mandatory": len(mandatory) - len(mandatory_provided),
            "missing_optional": total - len(mandatory) - (len(provided) - len(mandatory_provided)),
        }

    def export_manifest(
        self,
        manifest: Dict,
        output_format: str = "json",
        output_path: Optional[str] = None
    ) -> str:
        """
        Экспорт описи в указанном формате

        Args:
            manifest: Опись документов
            output_format: Формат вывода (json, csv, xlsx)
            output_path: Путь для сохранения (опционально)

        Returns:
            Путь к созданному файлу
        """
        manifest_id = manifest.get("manifest_id", "manifest")

        if output_path:
            output_file = Path(output_path)
        else:
            output_file = self.output_dir / f"{manifest_id}.{output_format}"

        if output_format == "json":
            return self._export_json(manifest, output_file)
        elif output_format == "csv":
            return self._export_csv(manifest, output_file)
        elif output_format == "xlsx":
            return self._export_xlsx(manifest, output_file)
        else:
            raise ValueError(f"Неподдерживаемый формат: {output_format}")

    def _export_json(self, manifest: Dict, output_file: Path) -> str:
        """Экспорт в JSON"""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)
        logger.info(f"Опись экспортирована в JSON: {output_file}")
        return str(output_file)

    def _export_csv(self, manifest: Dict, output_file: Path) -> str:
        """Экспорт в CSV"""
        items = manifest.get("items", [])

        if not items:
            logger.warning("Опись пуста, CSV не создан")
            return ""

        fieldnames = [
            "position", "document_id", "document_name", "type",
            "mandatory", "completion_status", "source_type",
            "format", "notes"
        ]

        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for item in items:
                writer.writerow(item)

        logger.info(f"Опись экспортирована в CSV: {output_file}")
        return str(output_file)

    def _export_xlsx(self, manifest: Dict, output_file: Path) -> str:
        """Экспорт в Excel (xlsx)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl не установлен, используется CSV")
            csv_file = output_file.with_suffix('.csv')
            return self._export_csv(manifest, csv_file)

        items = manifest.get("items", [])
        metrics = manifest.get("metrics", {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Опись документов"

        # Заголовки
        headers = [
            "№", "ID", "Наименование документа", "Тип", "Обязательный",
            "Статус", "Источник", "Формат", "Примечания"
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Данные
        status_colors = {
            self.STATUS_PROVIDED: "C6EFCE",
            self.STATUS_FROM_TEMPLATE: "C6EFCE",
            self.STATUS_MISSING: "FFC7CE",
            self.STATUS_NOT_PREPARED: "FFEB9C",
        }

        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.get("position"))
            ws.cell(row=row, column=2, value=item.get("document_id"))
            ws.cell(row=row, column=3, value=item.get("document_name"))
            ws.cell(row=row, column=4, value=item.get("type"))
            ws.cell(row=row, column=5, value="Да" if item.get("mandatory") else "Нет")
            ws.cell(row=row, column=6, value=item.get("completion_status"))
            ws.cell(row=row, column=7, value=item.get("source_type"))
            ws.cell(row=row, column=8, value=item.get("format"))
            ws.cell(row=row, column=9, value=item.get("notes"))

            # Цветовое кодирование статуса
            status = item.get("completion_status", "")
            if status in status_colors:
                fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = fill

        # Ширина колонок
        column_widths = [5, 15, 50, 15, 12, 15, 20, 15, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Добавляем сводку
        summary_row = len(items) + 3
        ws.cell(row=summary_row, column=1, value="СВОДКА").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value="Всего документов:")
        ws.cell(row=summary_row + 1, column=2, value=metrics.get("total_documents", 0))
        ws.cell(row=summary_row + 2, column=1, value="Готовность (%):")
        ws.cell(row=summary_row + 2, column=2, value=metrics.get("completeness_percentage", 0))

        wb.save(output_file)
        logger.info(f"Опись экспортирована в Excel: {output_file}")
        return str(output_file)

    def create_package_structure(self, package_id: str) -> Path:
        """
        Создание структуры каталогов пакета

        Args:
            package_id: Идентификатор пакета

        Returns:
            Путь к корневой директории пакета
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        package_dir = self.output_dir / f"Заявка_{package_id}_{timestamp}"

        # Создаем основную директорию
        package_dir.mkdir(parents=True, exist_ok=True)

        # Создаем подкаталоги
        for key, name in self.PACKAGE_STRUCTURE.items():
            subdir = package_dir / name
            subdir.mkdir(exist_ok=True)

        logger.info(f"Создана структура пакета: {package_dir}")
        return package_dir

    def copy_files_to_package(
        self,
        manifest: Dict,
        package_dir: Path,
        source_files: Dict[str, str]
    ) -> List[Dict]:
        """
        Копирование файлов в пакет

        Args:
            manifest: Опись документов
            package_dir: Директория пакета
            source_files: Словарь {document_id: file_path}

        Returns:
            Список информации о скопированных файлах
        """
        copied_files = []

        for item in manifest.get("items", []):
            doc_id = item.get("document_id")
            source_path = source_files.get(doc_id) or item.get("source_path")

            if not source_path or not Path(source_path).exists():
                continue

            # Определяем целевую директорию
            source_type = item.get("source_type", "")
            if source_type == "template_library":
                target_subdir = self.PACKAGE_STRUCTURE["templates"]
            elif item.get("type") == "form":
                target_subdir = self.PACKAGE_STRUCTURE["forms"]
            else:
                target_subdir = self.PACKAGE_STRUCTURE["user_docs"]

            target_dir = package_dir / target_subdir

            # Копируем файл
            source_file = Path(source_path)
            position = item.get("position", 0)
            target_file = target_dir / f"{position:02d}_{source_file.name}"

            try:
                shutil.copy2(source_file, target_file)
                copied_files.append({
                    "document_id": doc_id,
                    "source_path": str(source_file),
                    "target_path": str(target_file),
                    "file_size": target_file.stat().st_size,
                })
                logger.debug(f"Скопирован: {source_file.name} -> {target_file}")
            except Exception as e:
                logger.error(f"Ошибка копирования {source_file}: {e}")

        logger.info(f"Скопировано файлов: {len(copied_files)}")
        return copied_files

    def create_zip_archive(
        self,
        package_dir: Path,
        output_path: Optional[str] = None,
        password: Optional[str] = None
    ) -> str:
        """
        Упаковка пакета в ZIP-архив

        Args:
            package_dir: Директория пакета
            output_path: Путь к архиву (опционально)
            password: Пароль для архива (опционально, не реализовано)

        Returns:
            Путь к созданному архиву
        """
        if output_path:
            zip_path = Path(output_path)
        else:
            zip_path = package_dir.with_suffix('.zip')

        # Создаем архив
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in package_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(package_dir)
                    zipf.write(file_path, arcname)

        # Вычисляем контрольную сумму
        sha256_hash = hashlib.sha256()
        with open(zip_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)

        checksum = sha256_hash.hexdigest()

        # Сохраняем контрольную сумму
        checksum_file = zip_path.with_suffix('.sha256')
        with open(checksum_file, 'w') as f:
            f.write(f"{checksum}  {zip_path.name}\n")

        logger.info(f"Создан архив: {zip_path} (SHA256: {checksum[:16]}...)")
        return str(zip_path)

    def build_complete_package(
        self,
        manifest: Dict,
        source_files: Dict[str, str],
        procurement_id: str
    ) -> Dict[str, Any]:
        """
        Полное формирование пакета документов

        Args:
            manifest: Опись документов
            source_files: Словарь файлов
            procurement_id: ID закупки

        Returns:
            Информация о созданном пакете
        """
        # Создаем структуру
        package_dir = self.create_package_structure(procurement_id)

        # Копируем файлы
        copied_files = self.copy_files_to_package(manifest, package_dir, source_files)

        # Экспортируем опись в разные форматы
        manifest_json = self.export_manifest(manifest, "json", str(package_dir / "manifest.json"))
        manifest_csv = self.export_manifest(manifest, "csv", str(package_dir / "Опись_документов.csv"))

        try:
            manifest_xlsx = self.export_manifest(manifest, "xlsx", str(package_dir / "Опись_документов.xlsx"))
        except Exception:
            manifest_xlsx = None

        # Создаем архив
        zip_path = self.create_zip_archive(package_dir)

        return {
            "package_id": procurement_id,
            "package_dir": str(package_dir),
            "zip_path": zip_path,
            "manifest_files": {
                "json": manifest_json,
                "csv": manifest_csv,
                "xlsx": manifest_xlsx,
            },
            "copied_files_count": len(copied_files),
            "total_size_bytes": sum(f.get("file_size", 0) for f in copied_files),
            "created_at": datetime.now().isoformat(),
        }
